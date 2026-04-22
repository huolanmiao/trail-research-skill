"""Generate 风险预案 (Risk Prevention Plan) as .xlsx file."""

import sys
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Allow importing from the same directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_model import TrailResearchData, RiskItem


# Risk level color mapping
RISK_COLORS = {
    "安全": "92D050",       # Green
    "风险存在": "FFC000",    # Yellow/Gold
    "风险容易发生": "ED7D31", # Orange
    "危险": "FF0000",        # Red
    "很危险": "C00000",      # Dark Red
    "极度危险": "000000",    # Black
}

RISK_DESCRIPTIONS = {
    "安全": "没有环境风险",
    "风险存在": "没有环境风险，但有不可预见的风险存在",
    "风险容易发生": "环境风险较高，会发生高致伤率，需要特别重视此类风险发生",
    "危险": "风险发生概率低但致伤后果严重，需要特别重视此类风险发生",
    "很危险": "风险发生概率高，发生后严重容易致残",
    "极度危险": "风险无法承受，发生风险为概率高还存在容易致死",
}


def generate_risk_plan(data: TrailResearchData, output_path: str) -> str:
    """Generate the risk prevention plan Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "风险预案"

    # Styles
    title_font = Font(name='宋体', size=14, bold=True)
    header_font = Font(name='宋体', size=10.5, bold=True)
    cell_font = Font(name='宋体', size=10.5)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Column widths
    col_widths = {
        'A': 6,   # 序号
        'B': 18,  # 路线地名
        'C': 18,  # 风险原因
        'D': 12,  # 风险级别
        'E': 50,  # 应对要求
        'F': 35,  # 备注
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Title row (merged A1:F1)
    is_pku = "北大" in data.organization.unit_name or "徒协" in data.organization.unit_name
    if is_pku:
        title_text = f"{data.activity.date}徒协{data.route.name}徒步风险分析和应对要求"
    else:
        title_text = f"{data.activity.date}{data.route.name}徒步风险分析和应对要求"
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = title_text
    title_cell.font = title_font
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # Header row (row 2)
    headers = ["序号", "路线地名", "风险原因", "风险级别", "应对要求", "备注"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws.row_dimensions[2].height = 25

    # Risk rows
    for r, risk in enumerate(data.risks):
        row_num = r + 3
        values = [risk.seq, risk.location, risk.risk_cause, risk.risk_level, risk.mitigation, risk.notes]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if c == 4:  # Risk level column - center
                cell.alignment = center_align
                color = RISK_COLORS.get(risk.risk_level)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            elif c in (5, 6):
                cell.alignment = left_align
            else:
                cell.alignment = center_align
        ws.row_dimensions[row_num].height = 60

    # Color legend (columns I-J, offset to the right)
    legend_start_col = 9  # Column I
    legend_title_row = 1

    # Legend title
    legend_cell = ws.cell(row=legend_title_row, column=legend_start_col, value="色块解释参考")
    legend_cell.font = Font(name='宋体', size=10.5, bold=True)

    legend_items = list(RISK_COLORS.items())
    for i, (level, color) in enumerate(legend_items):
        row = i + 2
        # Color swatch
        swatch_cell = ws.cell(row=row, column=legend_start_col, value=level)
        swatch_cell.font = cell_font
        swatch_cell.alignment = center_align
        swatch_cell.border = thin_border
        swatch_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Description
        desc_cell = ws.cell(row=row, column=legend_start_col + 1, value=RISK_DESCRIPTIONS.get(level, ""))
        desc_cell.font = cell_font
        desc_cell.alignment = left_align
        desc_cell.border = thin_border

    ws.column_dimensions[get_column_letter(legend_start_col)].width = 12
    ws.column_dimensions[get_column_letter(legend_start_col + 1)].width = 40

    # Freeze panes
    ws.freeze_panes = 'A3'

    wb.save(output_path)
    return output_path


def build_default_risks(data: TrailResearchData) -> list[RiskItem]:
    """Build a default risk list based on route data. Used when no custom risks provided."""
    if data.risks:
        return data.risks

    route_name = data.route.name
    difficulty = data.route.difficulty_grade
    has_steep = any("陡坡" in h or "滑坠" in h or "暴露" in h for h in data.route.special_hazards)
    is_high_risk = difficulty in ("高级", "超高级")

    # Weather risk: infer season from date string
    weather_level = "安全"
    weather_mitigation = f"预计{data.activity.date}天气良好"
    weather_notes = "关注天气预报，如有降雨提醒大家准备雨衣；徒步过程中提醒队员及时增减衣物"

    date_str = data.activity.date
    month = None
    # Try "2026年4月22日" format
    if "月" in date_str:
        m = re.search(r'(\d+)月', date_str)
        if m:
            month = int(m.group(1))
    # Try "2026-04-22" or "2026/04/22" format
    elif "-" in date_str or "/" in date_str:
        sep = "-" if "-" in date_str else "/"
        parts = date_str.split(sep)
        if len(parts) >= 2:
            try:
                month = int(parts[1])
            except (ValueError, IndexError):
                pass

    if month is not None:
        if month in (6, 7, 8):
            weather_level = "风险存在"
            weather_mitigation = f"{month}月为夏季，需防范中暑和雷阵雨"
            weather_notes = "关注天气预报，携带充足饮水，准备防暑药品和雨衣"
        elif month in (12, 1, 2):
            weather_level = "风险存在"
            weather_mitigation = f"{month}月为冬季，需防范低温和冰雪"
            weather_notes = "关注天气预报，准备充足保暖衣物和备用厚手套袜子"

    # Exhaustion risk
    exhaustion_level = "风险存在"
    if is_high_risk:
        exhaustion_level = "危险"

    # Separation risk
    separation_level = "安全" if not is_high_risk else "风险存在"

    # Cannot reach destination
    dest_level = "安全" if data.route.bailout_routes else "风险容易发生"
    dest_mitigation = "徒步路线上下撤点密集" if data.route.bailout_routes else "做好时间管理，划出关门时间"

    # Steep slope
    steep_level = "安全"
    if has_steep:
        steep_level = "风险容易发生" if is_high_risk else "风险存在"
    steep_mitigation = "提醒队员穿越野鞋、徒步鞋等鞋底纹路较深的鞋防滑"
    if has_steep:
        steep_mitigation += "；提醒队员拉开间距"

    risks = [
        RiskItem(1, route_name, "交通意外", "安全",
                 "乘坐正规营运资质的交通工具",
                 "掌握包车司机及公司电话以备救援接驳用"),
        RiskItem(2, "", "下雨、失温、中暑", weather_level,
                 weather_mitigation, weather_notes),
        RiskItem(3, "", "队员体能耗竭，失去行动能力", exhaustion_level,
                 "行进过程中密切关注队员情况，选择合适的地点安排队伍休整，领队携带盐丸、高能量食品，应对队员可能出现的抽筋、低血糖等问题",
                 "提前预判，切勿等到发生耗竭再处理"),
        RiskItem(4, "", "因植被茂密、大雾等原因队伍间距过大走散", separation_level,
                 "收紧队伍，安排向导、押后；如需分队，前后队分别安排押后",
                 "所有队员知晓迷路后的应对预案"),
        RiskItem(5, "", "不能到达目的地", dest_level,
                 dest_mitigation,
                 "如发生意外，领队陪同下就近下撤"),
        RiskItem(6, "", "崴脚", "风险存在",
                 "碎石土路提醒队员注意脚下",
                 "禁止边走路边拍照"),
        RiskItem(7, "", "陡坡滑坠", steep_level,
                 steep_mitigation,
                 "根据领队之前带队该路线的经验，陡坡处风险可控"),
        RiskItem(8, "", "动植物风险（虫蛇和植物刺、划伤）", "风险存在",
                 "行进过程中注意脚下，如遇虫蛇注意避开；窄路提醒同学拉开间距注意防止树枝回弹伤人",
                 "该区域毒蛇、毒虫罕见"),
    ]

    # Add bail-out routes if available
    if data.route.bailout_routes:
        for i, br in enumerate(data.route.bailout_routes):
            risks.append(RiskItem(
                9 + i, "下撤路线", br,
                "安全",
                f"如需下撤，从{br}方向撤离",
                "领队熟悉下撤路线",
            ))

    return risks


def main():
    if len(sys.argv) < 3:
        print("Usage: python generate_risk_plan.py <input_json> <output_path>")
        sys.exit(1)

    input_json = sys.argv[1]
    output_path = sys.argv[2]

    try:
        data = TrailResearchData.from_json(input_json)
    except Exception as e:
        print(f"Error loading data from {input_json}: {e}", file=sys.stderr)
        sys.exit(1)

    if not data.risks:
        data.risks = build_default_risks(data)

    try:
        result = generate_risk_plan(data, output_path)
        print(f"Risk plan saved to: {result}")
    except Exception as e:
        print(f"Error generating risk plan: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
